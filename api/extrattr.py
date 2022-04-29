from logging import ERROR

import openpyxl
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response
from rest_framework.viewsets import GenericViewSet
from wechatpy.work import WeChatClient

from api.constants import IAC_SECRET, WX_SECRET
from api.serializers import UpdateAttrSerializer
from utils.logger import logger
from utils.wxiac import iac
from utils.wxmethod import wx_method


# Create your views here.
class MyExcelView(GenericViewSet):
    # url http://127.0.0.1:8000/v1/api/wx/extrattr/
    @action(methods=['post'], detail=False, url_path='extattr')
    def my_excel(self, request, *args, **kwargs) -> Response:
        # 0. 数据校验
        try:
            # 获取前端传入的请求体数据
            data = request.data
            # 创建序列化器进行反序列化
            serializer = UpdateAttrSerializer(data=data)
            # 调用序列化器的is_valid方法进行校验
            serializer.is_valid(raise_exception=True)
        except ValidationError as e:
            errors = f'UpdateAttr Validate Failed: {e}'
            logger.alert(level=ERROR, msg=errors)
            return Response({"result": "failed", "message": errors})
        domain: str = data['domain']
        excel = data['excel']

        # 1。获取token
        iac.get_access_token(IAC_SECRET.get('appid'), IAC_SECRET.get('secret'))

        """
        2.读取excel
        @param domain_p: 表单里的域账户信息
        @param field: 要修改的字段
        @param content: 修改的内容
        @param option: 执行的选项
        """
        wb = openpyxl.load_workbook(excel)
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])

        for row in range(2, ws.max_row + 1):
            domain_excel = ws.cell(row=row, column=2).value
            field = ws.cell(row=row, column=3).value
            content = ws.cell(row=row, column=4).value
            option = ws.cell(row=row, column=5).value

            # 判断field字段是否合法,判断执行者domain是否有权限
            wx_valid = wx_method.is_valid(field, domain, domain_excel)
            if wx_valid == 'not_field':
                return Response({'message': f'{domain_excel} 域账户的({field})字段不合法'}, status=status.HTTP_404_NOT_FOUND)
            elif wx_valid == 'net_domain':
                return Response({'message': f'执行用户：{domain} ，对{field}字段没有权限修改'}, status=status.HTTP_404_NOT_FOUND)

            if content == "清空":
                content = ""

            # 3.根据域账户获取wx-id
            """
            @param wx_id: 企业微信 id
            """
            wx_id = iac.get_user_info(domain=domain_excel)
            if not wx_id:
                logger.info(f'{domain_excel} :not wx_id ')
                continue

            """
            @param corp_id: 企业微信 corp
            @param secret: 密钥
            @param extattr_add: 企业微信个人信息
            """
            # 调用企业微信api获取用户信息
            client = WeChatClient(WX_SECRET.get('corp_id'), WX_SECRET.get('secret'))
            extattr_add = client.user.get(wx_id)['extattr']
            logger.info(f'update extattr {domain_excel} info: {extattr_add}')

            """
            @param option: 操作： 替换/更新
            @param extattr_add: 企业微信个人信息
            @param domain_excel: excel里的域账户
            @param field: 修改的字段
            @param content: 修改的内容

            方法：先根据 option 判断操作（替换/更新），再根据 field 判断要修改那个字段
            """
            extattr_add_update = wx_method.choice_option_field(option, extattr_add, domain_excel, field, content)

            try:
                # 调用企业微信api更改字段值
                client.user.update(user_id=wx_id, extattr=extattr_add_update)
                logger.info(f'{domain_excel} update success')
            except Exception as e:
                logger.info(f'{domain_excel} error: {e}')

        wb.close()
        return Response({'message': f'用户字段更新完成'}, status=status.HTTP_200_OK)
